from typing import Tuple

from openpyxl import load_workbook

from ..extensions import db
from ..models import Area


def import_areas_from_excel(file_like) -> Tuple[int, int]:
    """Import areas from first sheet. Accepts columns: Area / Nombre / name in A."""
    wb = load_workbook(file_like, data_only=True)
    ws = wb.active

    created = 0
    skipped = 0

    # Detect header
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row:
            continue
        if i == 1 and any(isinstance(x, str) and x.strip().lower() in {'area', 'nombre', 'name'} for x in row if x):
            continue

        name = None
        for cell in row:
            if isinstance(cell, str) and cell.strip():
                name = cell.strip()
                break
        if not name:
            continue

        if Area.query.filter_by(name=name).first():
            skipped += 1
            continue

        db.session.add(Area(name=name, is_active=True))
        created += 1

    return created, skipped
